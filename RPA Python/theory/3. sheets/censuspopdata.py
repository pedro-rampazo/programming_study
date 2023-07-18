import openpyxl
import os

os.chdir("/home/pedro/Development/git_space/programming_study/RPA Python/theory/3. sheets")

wb = openpyxl.load_workbook('censuspopdata.xlsx')
sheet = wb.get_sheet_by_name('Population by Census Tract')
file = open("censuspopdata.txt", "w")

county = set()

for c in range(sheet.max_row):
    county.add(sheet[f'C{c+1}'].value)

for y in county:
    census_tract = 0
    population = 0
    for row in tuple(sheet['A2':f'D{sheet.max_row}']):
        if row[2].value == y:
            census_tract += 1
            population += row[3].value
    file.write(f"{y} - Census Tract: {census_tract} | Total Population: {population}\n")

file.close()

# print(tuple(sheet['A1':'D3']))

# for row in tuple(sheet['A1':'D1']):
#     for sr in row:
#         print(sr.value)

