import openpyxl
import os

from openpyxl.utils import get_column_letter

os.chdir("/home/pedro/Development/git_space/programming_study/RPA Python/theory/3. sheets")

wb = openpyxl.load_workbook('example.xlsx')

print(type(wb))

print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('Sheet1')
# print(sheet['A1'])
# print(sheet['A1'].value)

# print(sheet['B1'])
# print(sheet['B1'].value)
# print(sheet['B1'].row)
# print(sheet['B1'].column)
# print(sheet['B1'].coordinate)

# print(sheet.cell(row=1, column=3))
# print(sheet.cell(row=1, column=3).value)

# print(sheet.max_row)
# print(sheet.min_row)

# print(get_column_letter(55)) # Transforma o número do argumento em letra da coluna

# print(tuple(sheet['A1':'C3']))

for row_of_cell_objects in sheet['A1':'C3']:
    for cell_obj in row_of_cell_objects:
        print(cell_obj.coordinate, cell_obj.value)
print('---END OF ROW---')

